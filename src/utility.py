import multiprocessing
import pickle
import os
import numpy as np
from openpyxl import load_workbook
import pandas as pd
from collections import Counter
import random
import math

cpus = int(multiprocessing.cpu_count())
min_fg_freq=float(1/100000)
max_bg_freq=float(1/150000)
min_tm=15
max_tm=45
max_gini=0.6
max_primer=500
min_amp_pred = 5
max_dimer_bp = 3
max_self_dimer_bp = 4
mismatch_penalty = 4
data_dir=os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data/project')

char_to_int_dict = {'A':0, 'G':1, 'T':2, 'C':3}

def create_pool(func, input_list, cpus):
    """Calculate the results of func applied to each value in input_list using a multiprocessing pool.

    Args:
        func (func): Function to apply to each value in the input list.
        input_list (list): List of values to which func will be applied.
        cpus (int): The number of processes to use.

    Returns:
        results: List of the results of function applied to the inputlist.
    """
    pool = multiprocessing.Pool(processes=cpus)
    results = pool.map(func, input_list)
    pool.close()
    return results

def flatten(l):
    """Flattens a multidimensional array.

    Args:
        l: The list to flatten.

    Returns:
        flattened_list: The flattened array.
    """

    return [item for sublist in l for item in sublist]

def mergeArrays(arr1, arr2):
    """Merges two sorted arrays, maintaining sorted order.

        Args:
            arr1: First array in sorted order.
            arr2: Second array in sorted order.

        Returns:
            arr3: The merged array in sorted order.
    """
    n1 = len(arr1)
    n2 = len(arr2)

    arr3 = []
    i = 0
    j = 0

    # Traverse both array
    while i < n1 and j < n2:
        if arr1[i] < arr2[j]:
            arr3.append(arr1[i])
            i = i + 1
        elif arr1[i] > arr2[j]:
            arr3.append(arr2[j])
            j = j + 1
        else:
            j = j + 1

    # Store remaining elements of first array
    while i < n1:
        arr3.append(arr1[i]);
        i = i + 1

    # Store remaining elements of second array
    while j < n2:
        arr3.append(arr2[j]);
        j = j + 1
    return arr3

def softmax(x):
    """Compute softmax values for each sets of scores in x.

        Args:
            x: Input value for softmax.

        Returns:
            s: Value of softmax function of x.
    """
    e_x = np.exp(x)
    return e_x / e_x.sum()

def sigmoid(x):
  return 1 / (1 + math.exp(-x))

def output_to_df(df, sheet_name, xls_path):
    """Output a pandas dataframe to a excel spreadsheet.

        Args:
            df: The pandas dataframe to be written to the spreadsheet.
            sheet_name: The sheet name of the excel file.
            xls_path: The path to the excel file.
    """
    book = load_workbook(xls_path)

    if sheet_name in book.sheetnames:
        sheet = book.get_sheet_by_name(sheet_name)
        book.remove(sheet)
    book.create_sheet(sheet_name)

    writer = pd.ExcelWriter(xls_path, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    df.to_excel(writer, sheet_name=sheet_name)
    writer.save()

def get_seq_length(genome):
    """
    Computes the sequence length of a genome in one fasta file.

    Args:
        genome: Fasta file of the genome.

    Returns:
        l: Total length of characters in the fasta file.
    """
    return sum(1 for _ in read_fasta_file(genome))

def get_all_seq_lengths(fname_genomes=None, cpus=8):
    """
    Computes the sequence length of multiple genomes.

    Args:
        fname_genomes: Fasta file of the genome.
        cpus: Number of cpus to be used.

    Returns:
        l: List of lengths of the individual genomes in all the fasta files.
    """
    # seq_lengths = create_pool(get_seq_length, fname_genomes, cpus)
    seq_lengths = []

    for fname_genome in fname_genomes:
        seq_lengths.append(get_seq_length(fname_genome))
    return seq_lengths

def longest_char_repeat(s, char):
    max_count = 0
    if s[0] == char:
        count = 1
    else:
        count = 0

    for i,c in enumerate(s):
        if i == 0:
            continue
        if c == char and s[i-1] == char:
            count += 1
        elif c == char:
            count = 1
        if count > max_count:
            max_count = count
    return max(count, max_count)

complement_dic ={'A':'T','T':'A','G':'C','C':'G', " ": " "}
def complement(text):
    return ''.join([complement_dic[c.upper()] if c.upper() in complement_dic else "N" for c in text])

def get_num_mismatches(x, y):
    num_mismatches = 0
    for i in range(len(x)):
        if x[i] != y[i]:
            num_mismatches += 1
    return num_mismatches

#returns the length of longest common substring, from wikipedia
def longest_common_substring(s1, s2):
    m = [[0] * (1 + len(s2)) for i in range(1 + len(s1))]
    longest, x_longest = 0, 0
    for x in range(1, 1 + len(s1)):
        for y in range(1, 1 + len(s2)):
            if s1[x - 1] == s2[y - 1]:
                m[x][y] = m[x - 1][y - 1] + 1
                if m[x][y] > longest:
                    longest = m[x][y]
                    x_longest = x
            else:
                m[x][y] = 0
    return longest

def read_fasta_file(fname):
    #exclude the first line
    with open(fname,'r') as f:
        header = f.readline()
        for line in f:
            for ch in ''.join(c for c in line.strip()):
                yield ch

def reverse(seq):
    return seq[::-1]

def reverse_complement(seq):
    alt_map = {'ins': '0'}
    for k, v in alt_map.items():
        seq = seq.replace(k, v)
    bases = list(seq)
    bases = reversed([complement_dic.get(base, base) for base in bases])
    bases = ''.join(bases)
    for k, v in alt_map.items():
        bases = bases.replace(v, k)
    return bases

def intersection(lst1, lst2):
    lst3 = [value for value in lst1 if value in lst2]
    return lst3

def gini_exact(array):
    """Calculate the Gini coefficient of a numpy array. Based on bottom equation from
    http://www.statsdirect.com/help/default.htm#nonparametric_methods/gini.htm
    All values are treated equally, arrays must be 1d

    Args:
        array (array): One-dimensional array or list of floats or ints.

    Returns:
        gini_index: The gini index of the array given.
    """
    if len(array) == 0:
        return 0
    array = np.array(array, dtype=float)
    if np.amin(array) < 0:
        # Values cannot be negative:
        array -= np.amin(array)
    # Values cannot be 0:
    array += 0.0000001
    # Values must be sorted:
    array = np.sort(array)
    # Index per array element:
    index = np.arange(1, array.shape[0] + 1)
    # Number of array elements:
    n = array.shape[0]
    # Gini coefficient:
    return ((np.sum((2 * index - n - 1) * array)) / (n * np.sum(array)))

def most_frequent(list):
    occurence_count = Counter(list)
    high_freq_primer = occurence_count.most_common(1)[0][0]

    high_freq_primers = []

    for k,v in occurence_count.items():
        if v == occurence_count[high_freq_primer]:
            high_freq_primers.append(k)

    return random.choice(high_freq_primers)

if __name__ == "__main__":
    # print(parameter.cpus)
    gini = gini_exact([4995, 14637, 18645, 20181, 29308, 35414, 37613, 38293, 40057, 45539, 47585, 49334, 56793, 58542, 60054, 63126, 80157, 84046, 85780, 86389, 90143, 94548, 96939, 99848, 99978, 102881, 112277, 115155, 121079, 123930, 126932, 137880, 140422, 151813, 171916, 176720, 178581, 184803, 185946, 190838, 193269, 198980, 199880, 202613, 203744, 205857, 218059, 229732, 230301, 242735, 245249, 246332, 246719, 247313, 251470, 252798, 262210, 276783, 279267, 282369, 294881, 300204, 306965, 307010, 312090, 318203, 320816, 322935, 323845, 326770, 329012, 331440, 334910, 350327, 356708, 356886, 363515, 365734, 366645, 375058, 386446, 387276, 390904, 393917, 405329, 407237, 408560, 408595, 416522, 416765, 419471, 421244, 422159, 431641, 435079, 441547, 444510, 445506, 447867, 448196, 449188, 450659, 454724, 458719, 459936, 474665, 474752, 476203, 482467, 486021, 496127, 499491, 516734, 528192, 550223, 552967, 564491, 568605, 571079, 578567, 580396, 586375, 596543, 596749, 603065, 611361, 614572, 619893, 624716, 629591, 629862, 630814, 638371, 641411, 643996, 651007, 657051, 668427, 673358, 685666, 687264, 688017, 689214, 693902, 697591, 702234, 710660, 717828, 723640, 725529, 726345, 728145, 729468, 734625, 738669, 741589, 744674, 749573, 754819, 787048, 787650, 790267, 798940, 829028, 833082, 833829, 838552, 844292, 845354, 849692, 850483, 853939, 854354, 860213, 871217, 875805, 877652, 877787, 878023, 883751, 885933, 892744, 895835, 897160, 900728, 902728, 908581, 913780, 914458, 922942, 939619, 945931, 953533, 959162, 974468, 974672, 975191, 988478, 995136, 998637, 999621, 1003960, 1004522, 1006942, 1009804, 1017841, 1025720, 1038664, 1039303, 1044660, 1052579, 1060631, 1084299, 1088970, 1103221, 1105652, 1108659, 1118219, 1123329, 1126773, 1129277, 1130431, 1136876, 1142641, 1143438, 1147934, 1148874, 1150238, 1152638, 1154677, 1157869, 1158957, 1160353, 1166108, 1170539, 1187944, 1193074, 1195235, 1200080, 1210146, 1222361, 1223072, 1223189, 1224915, 1230027, 1231756, 1233447, 1235267, 1236341, 1245374, 1245789, 1247994, 1266680, 1281239, 1283314, 1294666, 1296439, 1297315, 1299115, 1305397, 1310105, 1320460, 1320929, 1323805, 1332098, 1334711, 1354879, 1360108, 1362212, 1371891, 1376359, 1378129, 1383037, 1386645, 1388197, 1389840, 1393629, 1399859, 1401647, 1402296, 1405167, 1405194, 1406498, 1410358, 1414590, 1430890, 1432619, 1434902, 1436203, 1436221, 1440181, 1446019, 1455534, 1457348, 1460583, 1462656, 1469135, 1471958, 1476714, 1479733, 1490333, 1490732, 1491170, 1492557, 1503429, 1504402, 1507087, 1513530, 1519755, 1521568, 1524815, 1525257, 1545155, 1549745, 1555384, 1562042, 1567037, 1568424, 1570217, 1572445, 1586312, 1587254, 1590165, 1590348, 1597622, 1605328, 1608182, 1621165, 1623329, 1623437, 1627055, 1629591, 1635199, 1635238, 1643352, 1644831, 1645638, 1649139, 1650669, 1664070, 1667499, 1676806, 1686845, 1694248, 1694863, 1696883, 1708342, 1715245, 1718549, 1724567, 1724909, 1728367, 1733449, 1737872, 1741842, 1743895, 1748132, 1757870, 1761309, 1766600, 1773641, 1778093, 1779018, 1798717, 1799226, 1800384, 1805191, 1820191, 1821810, 1823220, 1824149, 1826740, 1829025, 1835582, 1841058, 1841739, 1843442, 1859081, 1862959, 1872224, 1886278, 1890858, 1891322, 1893437, 1898127, 1898288, 1901084, 1903344, 1927898, 1932277, 1934536, 1935113, 1938164, 1947408, 1951305, 1954515, 1961139, 1961567, 1961879, 1963981, 1964052, 1964417, 1964714, 1964960, 1965717, 1966447, 1968053, 1969522, 1974857, 1999434, 2010766, 2017081, 2030134, 2032207, 2038022, 2041138, 2041901, 2042381, 2049162, 2049198, 2053712, 2057317, 2063051, 2064507, 2068376, 2068436, 2071416, 2071988, 2073567, 2077809, 2079397, 2082860, 2093639, 2102089, 2105338, 2107666, 2107715, 2107963, 2110702, 2111486, 2112013, 2117966, 2122258, 2122390, 2126389, 2127100, 2133339, 2134986, 2144514, 2147070, 2152177, 2153213, 2154174, 2159885, 2162489, 2171820, 2172151, 2172400, 2181206, 2186638, 2188589, 2193143, 2194200, 2200088, 2200633, 2203327, 2209085, 2211464, 2220876, 2228455, 2229561, 2233335, 2234464, 2237812, 2238360, 2243677, 2253487, 2261416, 2265249, 2268466, 2268721, 2274617, 2277674, 2278433, 2281490, 2284104, 2293000, 2296406, 2300307, 2302496, 2306376, 2307430, 2307467, 2315759, 2315906, 2329856, 2334193, 2347013, 2351215, 2353241, 2358908, 2374184, 2375184, 2378273, 2386774, 2396812, 2398383, 2401510, 2411076, 2416359, 2416972, 2420243, 2422075, 2432008, 2434667, 2438967, 2450260, 2450320, 2452253, 2456138, 2456788, 2461687, 2463076, 2463578, 2468536, 2470271, 2476679, 2487544, 2493404, 2509307, 2513205, 2514468, 2523931, 2528925, 2530863, 2531163, 2540112, 2552799, 2553251, 2557396, 2563157, 2566574, 2575413, 2578887, 2579271, 2582179, 2584614, 2590101, 2591105, 2595529, 2599531, 2602619, 2603081, 2604434, 2605500, 2606849, 2612597, 2614156, 2619982, 2624648, 2627029, 2630176, 2631932, 2647117, 2648173, 2652367, 2653715, 2658825, 2660181, 2660352, 2663672, 2671074, 2682189, 2697004, 2699047, 2699666, 2699966, 2701246, 2701449, 2703331, 2708751, 2711242, 2711446, 2711995, 2722370, 2727384, 2748672, 2751537, 2754495, 2756045, 2757542, 2758265, 2786904, 2787223, 2792972, 2795022, 2800359, 2809860, 2811512, 2813554, 2817131, 2817449, 2819190, 2823334, 2823376, 2836879, 2846146, 2847716, 2848973, 2855565, 2859045, 2859702, 2860257, 2860908, 2873452, 2876867, 2881323, 2885807, 2890844, 2895409, 2895472, 2895826, 2904949, 2905097, 2906287, 2906758, 2908834, 2910379, 2910862, 2912583, 2913097, 2913223, 2923901, 2933077, 2935273, 2937287, 2940415, 2942992, 2949245, 2952438, 2952687, 2957871, 2959515, 2959527, 2961680, 2971340, 2971718, 2973909, 2976302, 2979844, 2986057, 2986550, 2989773, 2994391, 2995719, 3001009, 3004545, 3007599, 3008259, 3010491, 3015983, 3020265, 3029447, 3029882, 3031729, 3034536, 3039756, 3044540, 3058737, 3065184, 3065479, 3065775, 3066875, 3070145, 3079172, 3079225, 3090684, 3092037, 3092565, 3099834, 3103578, 3103893, 3106030, 3106117, 3114414, 3116724, 3119959, 3123527, 3124032, 3126875, 3129974, 3131949, 3132091, 3132115, 3138207, 3141225, 3141587, 3141980, 3142148, 3142816, 3146019, 3146194, 3150809, 3152663, 3155119, 3162117, 3163050, 3165064, 3170109, 3170448, 3171125, 3172542, 3174134, 3181013, 3181947, 3186076, 3189777, 3196284, 3196628, 3197053, 3202793, 3211227, 3214954, 3218725, 3220787, 3221323, 3226280, 3226394, 3230438, 3232141, 3234621, 3240923, 3247284, 3252698, 3258330, 3278203, 3282183, 3282901, 3288229, 3288559, 3293792, 3296030, 3296093, 3298976, 3300360, 3302879, 3309860, 3311848, 3318951, 3323041, 3329050, 3337867, 3345026, 3347294, 3350149, 3352923, 3354296, 3360147, 3369704, 3386776, 3389924, 3398872, 3402469, 3403698, 3405772, 3406806, 3410676, 3417829, 3424999, 3428985, 3431543, 3433977, 3439266, 3439497, 3443619, 3444663, 3445707, 3455053, 3463292, 3467894, 3474262, 3478800, 3482567, 3484396, 3488050, 3492607, 3493681, 3495463, 3496625, 3502678, 3515001, 3517369, 3518801, 3530323, 3531247, 3532440, 3532764, 3534639, 3536650, 3538297, 3559464, 3561535, 3562939, 3564224, 3570254, 3571983, 3577990, 3582449, 3590899, 3600411, 3612913, 3614222, 3621201, 3621588, 3622085, 3632841, 3633177, 3635422, 3637528, 3647168, 3656607, 3667253, 3671091, 3676900, 3684074, 3684118, 3685310, 3688432, 3697672, 3706332, 3714367, 3716172, 3721611, 3723203, 3731194, 3733747, 3751790, 3755790, 3759531, 3761295, 3763209, 3764871, 3776491, 3782854, 3794214, 3798616, 3818714, 3822515, 3829119, 3833166, 3835563, 3836974, 3837831, 3840464, 3852593, 3853827, 3856386, 3861882, 3862906, 3866562, 3869543, 3879408, 3884993, 3885469, 3888671, 3893592, 3894812, 3898192, 3900889, 3903401, 3903407, 3906750, 3906864, 3913569, 3915982, 3916465, 3918241, 3918529, 3925078, 3926737, 3945365, 3951097, 3954403, 3964004, 3966871, 3974607, 3974796, 3976476, 3978622, 3981122, 3982174, 3985179, 3985800, 3999251, 3999470, 3999569, 4000061, 4006093, 4008676, 4012312, 4016432, 4022976, 4025575, 4033355, 4034984, 4043961, 4053615, 4054095, 4066833, 4075847, 4077799, 4090761, 4091531, 4095882, 4098951, 4099152, 4099315, 4104251, 4105855, 4117582, 4117600, 4117657, 4122125, 4122185, 4126775, 4138427, 4145252, 4146523, 4149804, 4162063, 4162477, 4163086, 4167064, 4169069, 4173782, 4178543, 4184246, 4187719, 4189432, 4207194, 4208256, 4221494, 4226655, 4230664, 4233520, 4252410, 4255646, 4255667, 4256725, 4259041, 4265768, 4267061, 4276271, 4276948, 4283900, 4288449, 4288776, 4293112, 4299177, 4304484, 4308148, 4320051, 4321190, 4321412, 4323245, 4328548, 4345102, 4351455, 4351554, 4352925, 4356670, 4366359, 4372365, 4374501, 4390692, 4395385, 4396541, 4408730, 4410875, 4411369])
    print(gini)